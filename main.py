import os
import sys
from github import Github
import argparse
from envdefault import EnvDefault
import checks
from config_parser import Config
import openpyxl

config = Config("config.yml")
row = 1

def parse_arguments():
    """
    Parse them args
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('-t', '--token', action=EnvDefault, envvar='GIT_API_TOKEN', 
        help="GitHub token: set using -t or setting env var 'GIT_API_TOKEN'")
    parser.add_argument('-e', '--enterprise', action='store_true', help="Enable Github Enterprise")
    org_repo_group = parser.add_mutually_exclusive_group(required=True)
    org_repo_group.add_argument('-o', '--organization', help="Organization to scan all repos from")
    org_repo_group.add_argument('-r', '--repository', help="Single repo to scan PRs in")
    parser.add_argument('-b', '--branch', default='master', help="Base branch to use")
    parser.add_argument('-g', '--githost', required=False, action=EnvDefault, envvar='GIT_HOST', 
        help="GitHub enterprise host: set using -g or setting env var 'GIT_HOST'")
    parsed_args = parser.parse_args()
    if parsed_args.enterprise and not parsed_args.githost:
        parser.error("-e requires use of -g to set host, or env var 'GIT_HOST' to be set.")
        sys.exit()
    if parsed_args.repository and "/" not in parsed_args.repository:
        parser.error("Repo name should also contain organization name like: organization/repo_name")
        sys.exit()
    return parsed_args


def analyze_repo(repo, branch):
    is_valid = 0
    did_test = 0

    pulls = repo.get_pulls(state='closed', sort='created', direction='desc', base=branch)
    page_num = 0
    while is_valid <= config.constants["LOOKBACK"]:
        pr_page = pulls.get_page(page_num)
        if pr_page == []:
            print("No more PRs left to process")
            break
        for pr in pr_page:
            if is_valid == config.constants["LOOKBACK"]:
                break
            else:
                files_list = pr.get_files()
                if checks.pr_valid(files_list, config.ignore_list):
                    is_valid+=1
                    if checks.look_for_tests(files_list, config.test_pattern_list):
                        did_test+=1
        page_num += 1      
    
    write_spreadsheet(repo, is_valid, did_test)


def write_spreadsheet_headings():
    global row
    report = openpyxl.Workbook()
    sheet = report.active
    sheet.title = "Summary"
    sheet.cell(row=row, column=1, value="Repository")
    sheet.cell(row=row, column=2, value="Total PRs Checked")
    sheet.cell(row=row, column=3, value="Total Tested")
    sheet.cell(row=row, column=4, value="% of PRs Tested")
    report.save(config.constants["WORKSHEET"])
    row+=1


def write_spreadsheet(repo, num_analyzed, num_tested):
    global row
    report = openpyxl.load_workbook(config.constants["WORKSHEET"])
    sheet = report["Summary"]
    sheet.cell(row=row, column=1, value=repo.name)
    sheet.cell(row=row, column=2, value=num_analyzed)
    sheet.cell(row=row, column=3, value=num_tested)
    if num_analyzed > 0:
        percent = num_tested / num_analyzed * 100
    else:
        percent = 0
    sheet.cell(row=row, column=4, value=percent)
    report.save(config.constants["WORKSHEET"])
    row+=1


def main():
    args = parse_arguments()

    print("Logging in to Github...")
    if args.enterprise:
        # Enterprise
        base_url = f"https://{args.githost}/api/v3"
        g = Github(base_url=base_url, login_or_token=args.token)
    else:
        # Github
        g = Github(args.token)

    print("Writing output template...")
    write_spreadsheet_headings()

    repo = None
    if args.repository:
        print(f"Analyzing repository: {args.repository}")
        analyze_repo(g.get_repo(args.repository), args.branch)
    else:
        print(f"Analyzing org: {args.organization}")
        org = g.get_organization(args.organization)
        repos = org.get_repos()
        for repo in repos:
            print(f"Analyzing repository: {repo.name}")
            analyze_repo(repo, args.branch)


if __name__ == '__main__':
    main()